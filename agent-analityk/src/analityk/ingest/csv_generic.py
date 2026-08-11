"""Wczytywanie CSV/XLSX — docelowa, „porządna” droga wejścia danych.

PDF jest awaryjny: to wydruk, z którego trzeba odzyskiwać strukturę. Jeśli
CRM potrafi wyeksportować CSV/XLSX, używaj tego — parser jest odporniejszy,
a wgrywanie da się zautomatyzować.

Nagłówki dopasowywane są elastycznie (bez wielkości liter i ogonków), więc
zmiana nazwy kolumny w CRM nie wywraca wczytywania.
"""

from __future__ import annotations

import csv
import re
from datetime import datetime
from pathlib import Path

from ..models import Activity, klucz_pracownika, rozbij_adres, zbuduj_id

#: kanoniczna nazwa -> możliwe nagłówki w eksporcie
MAPOWANIE = {
    "data": ["data", "date", "data godzina", "datetime", "kiedy"],
    "pracownik": ["przypisany do", "pracownik", "agent", "assigned to", "opiekun", "user"],
    "typ": ["typ", "type", "rodzaj", "typ akcji"],
    "podtyp": ["mobilny", "podtyp", "kategoria", "subtype"],
    "powiazanie": ["zwiazany z", "związany z", "kontakt", "nieruchomosc", "obiekt", "related to"],
    "notatka": ["opis", "notatka", "note", "description", "komentarz", "uwagi"],
}

FORMATY_DATY = [
    "%d/%m/%Y %H:%M", "%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M",
    "%d/%m/%Y", "%d.%m.%Y", "%Y-%m-%d", "%m/%d/%Y %H:%M",
]

_OGONKI = str.maketrans("ąćęłńóśźż", "acelnoszz")


def _norm(s: str) -> str:
    return re.sub(r"\s+", " ", (s or "").strip().lower().translate(_OGONKI))


def _mapa_kolumn(naglowki: list[str]) -> dict[str, int]:
    znormalizowane = [_norm(h) for h in naglowki]
    mapa: dict[str, int] = {}
    for kanon, warianty in MAPOWANIE.items():
        for w in warianty:
            wn = _norm(w)
            for i, h in enumerate(znormalizowane):
                if h == wn or (wn in h and len(wn) > 3):
                    mapa[kanon] = i
                    break
            if kanon in mapa:
                break
    return mapa


def _parsuj_date(tekst: str) -> datetime | None:
    tekst = (tekst or "").strip()
    for fmt in FORMATY_DATY:
        try:
            return datetime.strptime(tekst, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(tekst)
    except ValueError:
        return None


def _kanal(typ: str, podtyp: str) -> str:
    t = _norm(f"{typ} {podtyp}")
    if "bezposredni" in t or "wizyta" in t or "drzwi" in t:
        return "bezposredni"
    if "telefon" in t or "polaczenie" in t or "tel" in t or "call" in t:
        return "telefon"
    if "mail" in t:
        return "email"
    if "spotkanie" in t or "prezentacja" in t:
        return "spotkanie"
    return "inny"


def _z_wierszy(wiersze: list[list[str]], nazwa_pliku: str) -> list[Activity]:
    if not wiersze:
        return []
    mapa = _mapa_kolumn(wiersze[0])
    brakujace = {"data", "pracownik"} - set(mapa)
    if brakujace:
        raise ValueError(
            f"W pliku brakuje kolumn: {', '.join(sorted(brakujace))}. "
            f"Znalezione nagłówki: {wiersze[0]}"
        )

    def pole(w: list[str], nazwa: str) -> str:
        i = mapa.get(nazwa)
        return (w[i] if i is not None and i < len(w) else "").strip()

    out: list[Activity] = []
    for w in wiersze[1:]:
        if not any(c.strip() for c in w):
            continue
        data = _parsuj_date(pole(w, "data"))
        nazwa_prac = re.sub(r"\s+", " ", pole(w, "pracownik")).strip()
        if data is None or not nazwa_prac:
            continue
        powiazanie = pole(w, "powiazanie")
        notatka = re.sub(r"\s+", " ", pole(w, "notatka")).strip()
        budynek, lokal = rozbij_adres(powiazanie, notatka)
        out.append(
            Activity(
                id=zbuduj_id(nazwa_prac, data, powiazanie, notatka),
                pracownik=klucz_pracownika(nazwa_prac),
                pracownik_nazwa=nazwa_prac,
                data=data,
                kanal=_kanal(pole(w, "typ"), pole(w, "podtyp")),
                podtyp=pole(w, "podtyp") or pole(w, "typ"),
                powiazanie=powiazanie,
                budynek=budynek,
                lokal=lokal,
                notatka=notatka,
                zrodlo=nazwa_pliku,
            )
        )
    return out


def wczytaj_csv(sciezka: str | Path) -> list[Activity]:
    sciezka = Path(sciezka)
    surowe = sciezka.read_text(encoding="utf-8-sig", errors="replace")
    proba = surowe[:4000]
    try:
        dialekt = csv.Sniffer().sniff(proba, delimiters=",;\t|")
    except csv.Error:
        dialekt = csv.excel
        dialekt.delimiter = ";" if proba.count(";") > proba.count(",") else ","
    wiersze = list(csv.reader(surowe.splitlines(), dialekt))
    return _z_wierszy(wiersze, sciezka.name)


def wczytaj_xlsx(sciezka: str | Path) -> list[Activity]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:  # pragma: no cover
        raise RuntimeError("Do plików XLSX potrzebny jest pakiet openpyxl") from e
    sciezka = Path(sciezka)
    ws = load_workbook(sciezka, read_only=True, data_only=True).active
    wiersze = [
        ["" if c is None else (c.strftime("%Y-%m-%d %H:%M")
                               if hasattr(c, "strftime") else str(c))
         for c in wiersz]
        for wiersz in ws.iter_rows(values_only=True)
    ]
    return _z_wierszy(wiersze, sciezka.name)
